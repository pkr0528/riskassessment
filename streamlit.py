import streamlit as st
import pandas as pd
import numpy as np
import joblib
import plotly.graph_objects as go
import subprocess
import json
import os
from pathlib import Path
import docx
from openpyxl import load_workbook
import google.generativeai as genai
import subprocess
from dashboard import render_dashboard
from dotenv import load_dotenv
load_dotenv()
st.set_page_config(page_title="CDP Score & Signal Predictor", layout="wide")

# =========================
# Document readers
# =========================

@st.cache_data
def read_docx(file_path):
    """Read DOCX file content"""
    try:
        doc = docx.Document(file_path)
        content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                content.append(paragraph.text)
        return "\n".join(content)
    except Exception as e:
        st.error(f"Error reading DOCX: {e}")
        return ""

@st.cache_data
def read_xlsx(file_path):
    """Read XLSX file content"""
    try:
        wb = load_workbook(file_path, read_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    content.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
        return "\n".join(content)
    except Exception as e:
        st.error(f"Error reading XLSX: {e}")
        return ""
 #gemini integration
try:
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
    if GEMINI_API_KEY:
        genai.configure(api_key=GEMINI_API_KEY)
    else:
        GEMINI_API_KEY = None
except Exception as e:
    st.error(f"Error configuring Gemini API: {e}")
    GEMINI_API_KEY = None

def analyze_with_gemini(prompt: str, model_name: str = "gemini-2.5-flash", max_output_tokens: int = 1024) -> str:

    """

    Call Gemini using the environment-configured key and return plain text.

    If an error occurs, returns string starting with 'Error'.

    """

    try:

        model = genai.GenerativeModel(model_name)

        response = model.generate_content(prompt)

        return response.text

    except Exception as e:

        return f"Error communicating with Gemini: {e}"


# =========================
# Persistent placeholders
# =========================

def get_progress_bar():
    if "_bar_slot" not in st.session_state:
        st.session_state["_bar_slot"] = st.empty()
        st.session_state["_bar"] = st.session_state["_bar_slot"].progress(0)
    return st.session_state["_bar"]

def reset_progress_bar():
    if "_bar_slot" in st.session_state:
        st.session_state["_bar_slot"].empty()
        del st.session_state["_bar_slot"]
    if "_bar" in st.session_state:
        del st.session_state["_bar"]

def ensure_flow_slot():
    if "_flow_slot" not in st.session_state:
        st.session_state["_flow_slot"] = st.empty()
    if "_flow_render_seq" not in st.session_state:
        st.session_state["_flow_render_seq"] = 0
    return st.session_state["_flow_slot"]

# =========================
# Stepper model + rendering
# =========================

STEP_NAMES = ["Load CSV", "Scale Features", "Predict Score", "Predict Signal", "Build Output"]

def init_steps():
    st.session_state["_steps"] = [{"name": n, "status": "pending"} for n in STEP_NAMES]

def current_index_from_statuses() -> int:
    statuses = [s["status"] for s in st.session_state.get("_steps", [])]
    for i, s in enumerate(statuses):
        if s == "running":
            return i
    for i, s in enumerate(statuses):
        if s == "pending":
            return i
    return len(statuses) - 1 if statuses else 0

def render_stepper(steps, current_idx: int) -> go.Figure:
    n = len(steps)
    xs = list(range(n))
    ys = [0] * n
    GREEN = "#22C55E"; BLUE = "#3B82F6"; GRAY = "#CBD5E1"
    fig = go.Figure()
    fig.add_shape(type="line", x0=0, y0=0, x1=n-1, y1=0, line=dict(color=GRAY, width=6))
    if n > 1 and current_idx > 0:
        fig.add_shape(type="line", x0=0, y0=0, x1=current_idx, y1=0, line=dict(color=GREEN, width=6))
    
    for i, label in enumerate(steps):
        color = GREEN if i < current_idx else BLUE if i == current_idx else GRAY
        fig.add_trace(go.Scatter(
            x=[xs[i]], y=[ys[i]], mode="markers+text",
            marker=dict(size=42, color=color, line=dict(width=2, color="white")),
            text=[label], textposition="top center",
            textfont=dict(size=12, color="#111827"), hoverinfo="text", showlegend=False
        ))
        status = st.session_state["_steps"][i]["status"]
        fig.add_annotation(
            x=xs[i], y=-0.35,
            text=("OK ✅" if status == "ok" else "RUNNING ⏳" if status == "running" else "PENDING •"),
            showarrow=False, font=dict(size=11, color="#111827")
        )
    
    fig.update_layout(
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-0.5, n-0.5]),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1, 1]),
        height=220, margin=dict(l=16, r=16, t=24, b=12),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        title=dict(text="Pipeline Progress", x=0.5, font=dict(size=18))
    )
    return fig

def draw_stepper():
    """Draw/update the stepper with a UNIQUE key each time."""
    flow_slot = ensure_flow_slot()
    st.session_state["_flow_render_seq"] += 1
    key = f"flow_chart_{st.session_state['_flow_render_seq']}"
    flow_slot.plotly_chart(
        render_stepper(STEP_NAMES, current_index_from_statuses()),
        use_container_width=True,
        key=key
    )

def set_step_status(step_name: str, status: str):
    for s in st.session_state["_steps"]:
        if s["name"] == step_name:
            s["status"] = status
            break
    draw_stepper()

# =========================
# Load models (cached)
# =========================

@st.cache_resource(show_spinner=False)
def load_models():
    reg = joblib.load("regression_model.pkl")
    clf = joblib.load("signal_classification_model.pkl")
    scl = joblib.load("standard_scaler_model.pkl")
    return reg, clf, scl

# =========================
# Main App
# =========================

st.title("📊 CDP Score & Signal Predictor")

# Create tabs
tab1, tab2,tab3= st.tabs(["🔮 Prediction", "📋 Analysis","📊 Dashboard"])

with tab1:
    # Original prediction functionality
    ensure_flow_slot()
    if "_steps" not in st.session_state:
        init_steps()
    draw_stepper()
    
    if "_bar" not in st.session_state:
        get_progress_bar()
    
    try:
        reg_model, clf_model, scaler = load_models()
    except Exception as e:
        st.error(f"❌ Model load failed: {e}")
        st.stop()
    
    uploaded_file = st.file_uploader("Upload cleaned CSV (with Partner_Code)", type=["csv"])
    
    if uploaded_file:
        # Fresh run: reset steps + progress bar + redraw once
        init_steps()
        reset_progress_bar()
        draw_stepper()
        bar = get_progress_bar()
        
        # Step 1: Load CSV
        try:
            set_step_status("Load CSV", "running")
            input_df = pd.read_csv(uploaded_file)
            id_cols = input_df[['Partner_Code']] if 'Partner_Code' in input_df.columns else pd.DataFrame()
            features_df = input_df.drop(columns=['Partner_Code'], errors='ignore')
            set_step_status("Load CSV", "ok"); bar.progress(20)
        except Exception as e:
            set_step_status("Load CSV", "fail")
            st.error(f"⚠️ CSV load failed: {e}"); st.stop()
        
        # Step 2: Scale Features
        try:
            set_step_status("Scale Features", "running")
            features_df_num = features_df.apply(pd.to_numeric, errors='coerce')
            med = features_df_num.median(numeric_only=True)
            features_df_num = features_df_num.fillna(med).fillna(0)
            X_scaled = scaler.transform(features_df_num)
            set_step_status("Scale Features", "ok"); bar.progress(40)
        except Exception as e:
            set_step_status("Scale Features", "fail")
            st.error(f"⚠️ Scaling failed: {e}"); st.stop()
        
        # Step 3: Predict Score
        try:
            set_step_status("Predict Score", "running")
            pred_score = reg_model.predict(X_scaled)
            set_step_status("Predict Score", "ok"); bar.progress(60)
        except Exception as e:
            set_step_status("Predict Score", "fail")
            st.error(f"⚠️ Score prediction failed: {e}"); st.stop()
        
        # Step 4: Predict Signal
        try:
            set_step_status("Predict Signal", "running")
            pred_signal = clf_model.predict(X_scaled)
            signal_map = {0: 'Green', 1: 'Red', 2: 'Yellow'}
            signal_labels = [signal_map.get(s, s) for s in pred_signal]
            set_step_status("Predict Signal", "ok"); bar.progress(80)
        except Exception as e:
            set_step_status("Predict Signal", "fail")
            st.error(f"⚠️ Signal prediction failed: {e}"); st.stop()
        
        # Step 5: Build Output
        try:
            set_step_status("Build Output", "running")
            results = id_cols.copy()
            results['Predicted Score'] = pred_score
            results['Predicted Signal'] = signal_labels
            
            # Store results in session state for analysis tab
            st.session_state['prediction_results'] = results
            st.session_state['input_features'] = features_df_num
            
            set_step_status("Build Output", "ok"); bar.progress(100)
            
            st.success("✅ Prediction Complete!")
            st.dataframe(results, use_container_width=True)
            st.download_button(
                "📥 Download Results",
                data=results.to_csv(index=False),
                file_name="cdp_predictions.csv",
                mime="text/csv"
            )
        except Exception as e:
            set_step_status("Build Output", "fail")
            st.error(f"⚠️ Building output failed: {e}")
    else:
        st.info("Upload a CSV to begin.")

with tab2:
    st.header("🔍 Signal Analysis & Reasoning")
    
    # Check if predictions are available
    if 'prediction_results' not in st.session_state:
        st.info("Please run predictions in the Prediction tab first.")
    else:
        # Partner selection for analysis
        results = st.session_state['prediction_results']
        if not results.empty:
            selected_partner = st.selectbox(
                "Select Partner for Analysis", 
                results['Partner_Code'].tolist()
            )
            
            if st.button("🔍 Analyze Signal Reasoning"):
                
                with st.spinner("Analyzing signal reasoning..."):
                        # Get partner data
                        partner_row = results[results['Partner_Code'] == selected_partner].iloc[0]
                        partner_features = st.session_state['input_features'][
                            st.session_state['input_features'].index == partner_row.name
                        ].iloc[0]
                        
                        # Read reference documents
                        docx_content = ""
                        xlsx_content = ""
                        
                        docx_path = "Partner Onboarding & Credit Limit Approval Matrix (1).docx"
                        xlsx_path = "CDP - Reasons for approval (2).xlsx"
                        
                        if os.path.exists(docx_path):
                            docx_content = read_docx(docx_path)
                        
                        if os.path.exists(xlsx_path):
                            xlsx_content = read_xlsx(xlsx_path)
                        
                        # Create analysis prompt
                        prompt = f"""
You are a credit risk analyst. Analyze the following partner's CDP signal prediction and provide reasoning.

Partner Code: {selected_partner}
Predicted Score: {partner_row['Predicted Score']:.2f}
Predicted Signal: {partner_row['Predicted Signal']}

Partner Features:
{partner_features.to_string()}

Reference Documentation:
=== Partner Onboarding & Credit Limit Approval Matrix ===
{docx_content[:2000] if docx_content else "Document not found"}

=== CDP Reasons for Approval ===
{xlsx_content[:2000] if xlsx_content else "Document not found"}

Please provide:
1. Key factors that influenced this signal prediction
2. Risk assessment based on the features
3. Recommendations for this partner
4. Reference to relevant approval criteria from the documents

Keep the analysis concise and actionable.
"""
                        
                        # Get analysis from Ollama
                        analysis = analyze_with_gemini(prompt)
                        
                        # Display results
                        col1, col2 = st.columns([1, 1])
                        
                        with col1:
                            st.subheader("📊 Partner Summary")
                            st.metric("Partner Code", selected_partner)
                            st.metric("Predicted Score", f"{partner_row['Predicted Score']:.2f}")
                            
                            signal_color = {
                                'Green': '🟢',
                                'Yellow': '🟡', 
                                'Red': '🔴'
                            }
                            st.metric(
                                "Predicted Signal", 
                                f"{signal_color.get(partner_row['Predicted Signal'], '⚪')} {partner_row['Predicted Signal']}"
                            )
                        
                        with col2:
                            st.subheader("🔍 Key Features")
                            # Show top 5 features with highest values
                            top_features = partner_features.abs().nlargest(5)
                            for feature, value in top_features.items():
                                st.write(f"**{feature}**: {value:.3f}")
                        
                        st.subheader("🧠 AI Analysis")
                        st.markdown(analysis)
                        
                        # Document status
                        st.subheader("📋 Reference Documents Status")
                        col1, col2 = st.columns(2)
                        with col1:
                            if docx_content:
                                st.success("✅ Partner Onboarding Matrix loaded")
                            else:
                                st.warning("⚠️ Partner Onboarding Matrix not found")
                        
                        with col2:
                            if xlsx_content:
                                st.success("✅ CDP Reasons document loaded")
                            else:
                                st.warning("⚠️ CDP Reasons document not found")

        # Batch analysis option
        st.markdown("---")
        st.subheader("📈 Batch Analysis")
        confirm_batch = st.checkbox("I understand this will call Gemini for each partner in the batch and may incur costs.")
        if st.button("🔄 Analyze All Red/Yellow Signals"):
            if not GEMINI_API_KEY:

                st.error("❌ No Gemini API key found in environment (GEMINI_API_KEY).")

                st.stop()
            else:
                high_risk = results[results['Predicted Signal'].isin(['Red', 'Yellow'])]
                if len(high_risk) > 0:
                    st.write(f"Found {len(high_risk)} partners with Red/Yellow signals:")
                    
                    analysis_results = []
                    progress_bar = st.progress(0)
                    
                    for i, (idx, partner) in enumerate(high_risk.iterrows()):
                        progress_bar.progress((i + 1) / len(high_risk))
                        
                        # Simple batch analysis prompt
                        prompt = f"""
As a credit risk analyst, briefly explain why Partner {partner['Partner_Code']} received a {partner['Predicted Signal']} signal with score {partner['Predicted Score']:.2f}.

Provide exactly 3 bullet points:
• Primary risk factor
• Secondary concern  
• Recommended action

Keep each point under 25 words.
"""
                        
                        analyze_with_gemini(prompt)
                        analysis_results.append({
                            'Partner_Code': partner['Partner_Code'],
                            'Signal': partner['Predicted Signal'],
                            'Score': partner['Predicted Score'],
                            'Analysis': analysis[:200] + "..." if len(analysis) > 200 else analysis
                        })
                    
                    # Display batch results
                    batch_df = pd.DataFrame(analysis_results)
                    st.dataframe(batch_df, use_container_width=True)
                else:
                    st.info("No Red or Yellow signals found in predictions.")


# after you run predictions and have:
# st.session_state['prediction_results']  (DataFrame)
# st.session_state['input_features']     (DataFrame)
# and you already have analyze_with_gemini(prompt, model_name) function 

with tab3:
    # Run dashboard independently - it will let users upload CSV
    # Optionally pass analyze_fn if you have analyze_with_gemini defined in your main file:
    try:
        analyze_fn = analyze_with_gemini  # if defined in your streamlit.py (Gemini)
    except NameError:
        analyze_fn = None

   
    
    render_dashboard(
        prediction_results=st.session_state.get('prediction_results'),      # preferred
        input_features=st.session_state.get('input_features'),              # optional
        analyze_fn=analyze_with_gemini,                                     # your existing analyze function
        gemini_model_name="gemini-2.5-flash",
        use_upload=False   # <- important: no uploader shown in Tab3
    )
