
import streamlit as st
import pandas as pd
import numpy as np
import joblib
import plotly.graph_objects as go
import subprocess
import json
import os
from pathlib import Path
import docx
from openpyxl import load_workbook
import google.generativeai as genai
from dotenv import load_dotenv

from cdp_utils import aggregate_transactional_features, apply_final_scoring
from dashboard import render_dashboard
load_dotenv(dotenv_path=".env")
st.set_page_config(page_title="CDP Score & Signal Predictor", layout="wide")

# =========================
# Document readers
# =========================

@st.cache_data
def read_docx(file_path):
    """Read DOCX file content"""
    try:
        doc = docx.Document(file_path)
        content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                content.append(paragraph.text)
        return "\n".join(content)
    except Exception as e:
        st.error(f"Error reading DOCX: {e}")
        return ""

@st.cache_data
def read_xlsx(file_path):
    """Read XLSX file content"""
    try:
        wb = load_workbook(file_path, read_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    content.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
        return "\n".join(content)
    except Exception as e:
        st.error(f"Error reading XLSX: {e}")
        return ""
 #gemini integration
ENV_GEMINI_KEY = os.getenv("GEMINI_API_KEY")

if ENV_GEMINI_KEY:
    try:
        genai.configure(api_key=ENV_GEMINI_KEY)
        _GEMINI_AVAILABLE = True
    except Exception as e:
        st.error(f"Failed to configure Gemini API: {e}")
        _GEMINI_AVAILABLE = False
else:
    _GEMINI_AVAILABLE = False

def analyze_with_gemini(prompt: str, model_name: str = "gemini-2.5-flash", max_output_tokens: int = 1024) -> str:
    """
    Safe wrapper for calling Gemini.
    - If GEMINI_API_KEY not set, return friendly message.
    - Otherwise call Gemini and return text.
    """
    if not _GEMINI_AVAILABLE:
        return "AI not configured. Please set GEMINI_API_KEY in your environment."

    try:
        model = genai.GenerativeModel(model_name)
        response = model.generate_content(prompt)
        return getattr(response, "text", str(response))
    except Exception as e:
        return f"AI call failed: {str(e)}"


# =========================
# Persistent placeholders
# =========================

def get_progress_bar():
    if "_bar_slot" not in st.session_state:
        st.session_state["_bar_slot"] = st.empty()
        st.session_state["_bar"] = st.session_state["_bar_slot"].progress(0)
    return st.session_state["_bar"]

def reset_progress_bar():
    if "_bar_slot" in st.session_state:
        st.session_state["_bar_slot"].empty()
        del st.session_state["_bar_slot"]
    if "_bar" in st.session_state:
        del st.session_state["_bar"]

def ensure_flow_slot():
    if "_flow_slot" not in st.session_state:
        st.session_state["_flow_slot"] = st.empty()
    if "_flow_render_seq" not in st.session_state:
        st.session_state["_flow_render_seq"] = 0
    return st.session_state["_flow_slot"]

# =========================
# Stepper model + rendering
# =========================

STEP_NAMES = ["Load CSV", "Scale Features", "Predict Score", "Predict Signal", "Build Output"]

def init_steps():
    st.session_state["_steps"] = [{"name": n, "status": "pending"} for n in STEP_NAMES]

def current_index_from_statuses() -> int:
    statuses = [s["status"] for s in st.session_state.get("_steps", [])]
    for i, s in enumerate(statuses):
        if s == "running":
            return i
    for i, s in enumerate(statuses):
        if s == "pending":
            return i
    return len(statuses) - 1 if statuses else 0

def render_stepper(steps, current_idx: int) -> go.Figure:
    n = len(steps)
    xs = list(range(n))
    ys = [0] * n
    GREEN = "#22C55E"; BLUE = "#3B82F6"; GRAY = "#CBD5E1"
    fig = go.Figure()
    fig.add_shape(type="line", x0=0, y0=0, x1=n-1, y1=0, line=dict(color=GRAY, width=6))
    if n > 1 and current_idx > 0:
        fig.add_shape(type="line", x0=0, y0=0, x1=current_idx, y1=0, line=dict(color=GREEN, width=6))
    
    for i, label in enumerate(steps):
        color = GREEN if i < current_idx else BLUE if i == current_idx else GRAY
        fig.add_trace(go.Scatter(
            x=[xs[i]], y=[ys[i]], mode="markers+text",
            marker=dict(size=42, color=color, line=dict(width=2, color="white")),
            text=[label], textposition="top center",
            textfont=dict(size=12, color="#111827"), hoverinfo="text", showlegend=False
        ))
        status = st.session_state["_steps"][i]["status"]
        fig.add_annotation(
            x=xs[i], y=-0.35,
            text=("OK ✅" if status == "ok" else "RUNNING ⏳" if status == "running" else "PENDING •"),
            showarrow=False, font=dict(size=11, color="#111827")
        )
    
    fig.update_layout(
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-0.5, n-0.5]),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1, 1]),
        height=220, margin=dict(l=16, r=16, t=24, b=12),
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        title=dict(text="Pipeline Progress", x=0.5, font=dict(size=18))
    )
    return fig

def draw_stepper():
    """Draw/update the stepper with a UNIQUE key each time."""
    flow_slot = ensure_flow_slot()
    st.session_state["_flow_render_seq"] += 1
    key = f"flow_chart_{st.session_state['_flow_render_seq']}"
    flow_slot.plotly_chart(
        render_stepper(STEP_NAMES, current_index_from_statuses()),
        use_container_width=True,
        key=key
    )

def set_step_status(step_name: str, status: str):
    for s in st.session_state["_steps"]:
        if s["name"] == step_name:
            s["status"] = status
            break
    draw_stepper()

# =========================
# Load models (cached)
# =========================

@st.cache_resource(show_spinner=False)
def load_models():
    reg = joblib.load("regression_model.pkl")
    clf = joblib.load("signal_classification_model.pkl")
    scl = joblib.load("standard_scaler_model.pkl")
    return reg, clf, scl

# =========================
# Main App
# =========================

st.title("CDP Score & Signal Predictor")

# Create tabs
tab1, tab2,tab3= st.tabs(["Prediction", "Analysis","Dashboard"])

with tab1:
    # Original prediction functionality
    ensure_flow_slot()
    if "_steps" not in st.session_state:
        init_steps()
    draw_stepper()

    if "_bar" not in st.session_state:
        get_progress_bar()

    try:
        reg_model, clf_model, scaler = load_models()
    except Exception as e:
        st.error(f"Model load failed: {e}")
        st.stop()

    uploaded_file = st.file_uploader("Upload cleaned CSV (with Partner_Code)", type=["csv"])

    if uploaded_file:
        # Fresh run: reset steps + progress bar + redraw once
        init_steps()
        reset_progress_bar()
        draw_stepper()
        bar = get_progress_bar()

        # Step 1: Load CSV
        try:
            set_step_status("Load CSV", "running")
            input_df = pd.read_csv(uploaded_file)
            id_cols = input_df[['Partner_Code']] if 'Partner_Code' in input_df.columns else pd.DataFrame()
            features_df = input_df.drop(columns=['Partner_Code'], errors='ignore')
            set_step_status("Load CSV", "ok"); bar.progress(20)
        except Exception as e:
            set_step_status("Load CSV", "fail")
            st.error(f"CSV load failed: {e}"); st.stop()

        # Step 2: Scale Features
        try:
            set_step_status("Scale Features", "running")
            features_df_num = features_df.apply(pd.to_numeric, errors='coerce')
            med = features_df_num.median(numeric_only=True)
            features_df_num = features_df_num.fillna(med).fillna(0)
            X_scaled = scaler.transform(features_df_num)
            set_step_status("Scale Features", "ok"); bar.progress(40)
        except Exception as e:
            set_step_status("Scale Features", "fail")
            st.error(f" Scaling failed: {e}"); st.stop()

        # Step 3: Predict Score
        try:
            set_step_status("Predict Score", "running")
            pred_score = reg_model.predict(X_scaled)
            set_step_status("Predict Score", "ok"); bar.progress(60)
        except Exception as e:
            set_step_status("Predict Score", "fail")
            st.error(f" Score prediction failed: {e}"); st.stop()

        # Step 4: Predict Signal
        try:
            set_step_status("Predict Signal", "running")
            pred_signal = clf_model.predict(X_scaled)
            signal_map = {0: 'Green', 1: 'Red', 2: 'Yellow'}
            signal_labels = [signal_map.get(s, s) for s in pred_signal]
            set_step_status("Predict Signal", "ok"); bar.progress(80)
        except Exception as e:
            set_step_status("Predict Signal", "fail")
            st.error(f" Signal prediction failed: {e}"); st.stop()

        # Step 5: Build Output
        try:
            set_step_status("Build Output", "running")
            if 'Partner_Code' not in input_df.columns:
                st.error(" Uploaded CSV must contain Partner_Code column.")
                st.stop()
            results = pd.DataFrame({'Partner_Code': input_df['Partner_Code']})
            results['Predicted Score'] = pred_score
            results['Predicted Signal'] = signal_labels
            # === NEW: transactional enrichment ===
            from cdp_utils import aggregate_transactional_features, apply_final_scoring

            txn_df = pd.read_csv("partner_transactional_dataset.csv")
            agg_df = aggregate_transactional_features(txn_df,partner_col="Partner_Code")
            


            final_results = apply_final_scoring(
                results_df=results,
                agg_df=agg_df,
                partner_key="Partner_Code",
                alpha=0.7
            )

            # Store enriched results in session state
            st.session_state['prediction_results'] = final_results
            st.session_state['input_features'] = agg_df

            set_step_status("Build Output", "ok"); bar.progress(100)

            st.success(" Prediction Complete!")
            st.dataframe(
                final_results[['Partner_Code', 'Predicted Score', 'Predicted Signal', 'Final Score', 'Final Signal']],
                use_container_width=True
            )
            st.download_button(
                "Download Results",
                data=final_results.to_csv(index=False),
                file_name="cdp_predictions.csv",
                mime="text/csv"
            )
        except Exception as e:
            set_step_status("Build Output", "fail")
            st.error(f"⚠️ Building output failed: {e}")
    else:
        st.info("Upload a CSV to begin.")
# ------------------------------
# Tab 2: Signal Analysis & Reasoning
# ------------------------------
with tab2:
    st.header("🔍 Signal Analysis & Reasoning")

    # Ensure predictions exist
    if 'prediction_results' not in st.session_state or st.session_state['prediction_results'] is None:
        st.info("Please run predictions in the Prediction tab first.")
    else:
        results = st.session_state['prediction_results']
        if results.empty:
            st.info("No prediction results found. Please run the Prediction tab.")
        else:
            # prefer final columns if available
            score_col = "Final Score" if "Final Score" in results.columns else "Predicted Score"
            signal_col = "Final Signal" if "Final Signal" in results.columns else "Predicted Signal"

            # partner selector
            partner_list = results['Partner_Code'].tolist() if 'Partner_Code' in results.columns else results.index.astype(str).tolist()
            selected_partner = st.selectbox("Select Partner for Analysis", partner_list)

            def _get_partner_row(results_df, partner_code):
                """Return partner row Series (robust)."""
                try:
                    return results_df[results_df['Partner_Code'] == partner_code].iloc[0]
                except Exception:
                    # fallback: try index match
                    try:
                        idx = results_df.index[results_df.index.astype(str) == partner_code][0]
                        return results_df.loc[idx]
                    except Exception:
                        return results_df.iloc[0]

            def _get_partner_kpis(feats_df, partner_code):
                """Return partner-level KPI Series if available, else empty Series."""
                if feats_df is None:
                    return pd.Series(dtype="float64")
                try:
                    if 'Partner_Code' in feats_df.columns:
                        tmp = feats_df[feats_df['Partner_Code'] == partner_code]
                        if not tmp.empty:
                            return tmp.iloc[0]
                    # otherwise assume feats_df is indexed in same order as results; attempt ordinal alignment
                    try:
                        ridx = results[results['Partner_Code'] == partner_code].index[0]
                        return feats_df.reset_index(drop=True).iloc[ridx]
                    except Exception:
                        return pd.Series(dtype="float64")
                except Exception:
                    return pd.Series(dtype="float64")

            def _get_model_input_row():
                """
                Try a few common session_state keys for raw model input rows.
                If you stored original input under a name, add it here.
                """
                for key in ("raw_input", "input_raw", "cdp_input", "input_features_raw"):
                    if key in st.session_state and st.session_state[key] is not None:
                        df = st.session_state[key]
                        # try to match by Partner_Code
                        try:
                            if 'Partner_Code' in df.columns:
                                tmp = df[df['Partner_Code'] == selected_partner]
                                if not tmp.empty:
                                    return tmp.iloc[0]
                        except Exception:
                            pass
                return pd.Series(dtype="float64")

            # When user clicks analyze
            if st.button("🔍 Analyze Signal Reasoning"):
                with st.spinner("Analyzing signal reasoning..."):
                    try:
                        prow = _get_partner_row(results, selected_partner)
                        kpis = _get_partner_kpis(st.session_state.get('input_features'), selected_partner)
                        model_input = _get_model_input_row()

                        # choose KPI keys to include (order matters: highest value to show abnormal signals)
                        candidate_kpis = [
                            "avg_days_past_due","max_days_past_due","late_ratio",
                            "dispute_ratio","adjustment_ratio","allocation_ratio","collection_rate",
                            "total_invoices","total_installments","total_invoice_amount",
                        ]
                        # filter to existing columns and present values
                        kpi_lines = []
                        for k in candidate_kpis:
                            if k in kpis.index:
                                val = kpis.get(k)
                                kpi_lines.append(f"- {k}: {val}")
                        if not kpi_lines:
                            kpi_lines = ["- (No transactional KPI snapshot available)"]

                        # also include top model input numeric features (if available)
                        input_lines = []
                        if isinstance(model_input, pd.Series) and not model_input.empty:
                            try:
                                numeric_input = pd.to_numeric(model_input, errors="coerce").dropna()
                                if not numeric_input.empty:
                                    top_input = numeric_input.abs().nlargest(6)
                                    for nm, v in top_input.items():
                                        input_lines.append(f"- {nm}: {v}")
                            except Exception:
                                pass
                        if not input_lines:
                            input_lines = ["- (No raw CDP input snapshot available)"]

                        # Build compact prompt
                        final_score_str = f"{prow.get('Final Score', prow.get('Predicted Score', 'N/A')):.2f}" if pd.notna(prow.get('Final Score', prow.get('Predicted Score', np.nan))) else "N/A"
                        final_signal_str = prow.get('Final Signal', prow.get('Predicted Signal', 'N/A'))

                        prompt = f"""
You are a senior credit risk analyst. Provide a short, structured analysis for Partner {selected_partner}.

Context:
- Final Score: {final_score_str}
- Final Signal: {final_signal_str}
- Model Predicted Score: {prow.get('Predicted Score', 'N/A')}
- Model Predicted Signal: {prow.get('Predicted Signal', 'N/A')}
- Override reason: {prow.get('OverrideReason', 'None')}
- Adjustment delta: {prow.get('AdjustmentDelta', 'N/A')}

Top transactional KPIs (most relevant):
{chr(10).join(kpi_lines)}

Top raw CDP input features (recent / largest):
{chr(10).join(input_lines)}

Requirements:
1) Header: Top 3 Key Factors (bullets).
2) Risk Assessment: 1-2 sentences explaining likelihood / urgency.
3) Recommendation: Approve / Review / Decline with one-line justification.
4) Quick Action Plan: 3 prioritized steps (1 line each).

Return only Markdown with headings and bullets and keep it concise (max 200-300 words).
"""

                        # Call AI if available
                        if analyze_with_gemini is None:
                            ai_text = "AI analysis not configured in this environment."
                        else:
                            try:
                                ai_text = analyze_with_gemini(prompt)
                            except Exception as e:
                                ai_text = f"Error calling AI: {e}"

                        # UI rendering
                        col1, col2 = st.columns([1,1])
                        with col1:
                            st.subheader("Partner Summary")
                            st.metric("Partner Code", selected_partner)
                            fs_val = prow.get("Final Score", prow.get("Predicted Score", np.nan))
                            st.metric("Final Score", f"{fs_val:.2f}" if pd.notna(fs_val) else "N/A")
                            sig_val = prow.get("Final Signal", prow.get("Predicted Signal", "N/A"))
                            signal_color = {'Green': '🟢', 'Yellow': '🟡', 'Red': '🔴'}
                            st.metric("Final Signal", f"{signal_color.get(sig_val,'⚪')} {sig_val}")
                            # small KPI snippet
                            st.markdown("**Top transactional KPIs (snapshot)**")
                            for line in kpi_lines[:6]:
                                st.write(line)

                        with col2:
                            st.subheader(" Top Model Input Features")
                            for line in input_lines[:6]:
                                st.write(line)

                        st.markdown("###  AI Analysis")
                        st.markdown(ai_text)

                    except Exception as e:
                        st.error(f"Error while preparing analysis: {e}")

            # Batch analysis option (careful: may incur API calls)
            st.markdown("---")
            st.subheader("Batch Analysis")
            confirm_batch = st.checkbox("I understand this will call the AI for each partner and may incur costs.")
            run_batch = st.button(" Analyze All Red/Yellow Signals")
            if run_batch:
                if not confirm_batch:
                    st.info("Please confirm the checkbox before running batch analysis.")
                
                else:
                    signal_col = "Final Signal" if "Final Signal" in results.columns else "Predicted Signal"
                    score_col = "Final Score" if "Final Score" in results.columns else "Predicted Score"
                    high_risk = results[results[signal_col].isin(['Red','Yellow'])]
                    if high_risk.empty:
                        st.info("No Red/Yellow partners found.")
                    else:
                        progress = st.progress(0)
                        batch_out = []
                        for i, (_, partner) in enumerate(high_risk.iterrows()):
                            try:
                                partner_code = partner.get('Partner_Code', str(partner.name))
                                prow = _get_partner_row(results, partner_code)
                                kpis = _get_partner_kpis(st.session_state.get('input_features'), partner_code)
                                # build short prompt (reuse same logic, but minimal to save tokens)
                                kp_lines = []
                                for k in ["late_ratio","avg_days_past_due","max_days_past_due","dispute_ratio","collection_rate"]:
                                    if k in kpis.index:
                                        kp_lines.append(f"- {k}: {kpis.get(k)}")
                                if not kp_lines:
                                    kp_lines = ["- (no kpi)"]
                                prompt = f"Partner {partner_code} | Signal: {prow.get(signal_col)} | Score: {prow.get(score_col)}\nTop KPIs:\n" + "\n".join(kp_lines) + "\nProvide 3 bullets: primary risk, secondary concern, recommended action. Keep each <25 words."
                                ai_resp = analyze_with_gemini(prompt)
                                batch_out.append({
                                    "Partner_Code": partner_code,
                                    "Signal": prow.get(signal_col),
                                    "Score": prow.get(score_col),
                                    "Analysis": ai_resp if isinstance(ai_resp, str) else str(ai_resp)
                                })
                            except Exception as e:
                                batch_out.append({
                                    "Partner_Code": partner.get("Partner_Code", str(partner.name)),
                                    "Signal": partner.get(signal_col),
                                    "Score": partner.get(score_col),
                                    "Analysis": f"Error: {e}"
                                })
                            progress.progress((i+1)/len(high_risk))
                        st.dataframe(pd.DataFrame(batch_out), use_container_width=True)


with tab3:
    
    render_dashboard(
        prediction_results=st.session_state.get('prediction_results'),      # preferred
        input_features=st.session_state.get('input_features'),              # optional
        analyze_fn=analyze_with_gemini,                                     # your existing analyze function
        gemini_model_name="gemini-2.5-flash",
        use_upload=False   # <- important: no uploader shown in Tab3
    )
