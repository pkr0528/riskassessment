import sys
from pathlib import Path

parent_dir = Path(__file__).parent.parent
sys.path.insert(0, str(parent_dir))

import streamlit as st
import pandas as pd
import numpy as np
import joblib
import plotly.graph_objects as go

import docx
from openpyxl import load_workbook
from dotenv import load_dotenv

from app.core.config import settings
from app.services.dashboard_service import render_dashboard
from app.services.risk_analyser import analyze_with_gemini
from app.utils.utils import STEP_NAMES

# Load Env file
load_dotenv()

st.set_page_config(page_title="CDP Score & Signal Predictor", layout="wide")

# Custom CSS for Navigation Bar Style Tabs
st.markdown(
    """
<style>
    /* Hide default streamlit style */
    .stTabs [data-baseweb="tab-list"] {
        background-color: #2c3e50;
        border-radius: 10px 10px 0 0;
        padding: 5px;
        margin-bottom: 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }

    /* Style individual tabs */
    .stTabs [data-baseweb="tab"] {
        height: 60px;
        padding: 0 24px;
        background-color: transparent;
        border-radius: 8px;
        margin: 0 4px;
        font-weight: 600;
        font-size: 18px;
        color: #ecf0f1;
        border: none;
        transition: all 0.3s ease;
    }

    /* Hover effect */
    .stTabs [data-baseweb="tab"]:hover {
        background-color: #34495e;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }

    /* Active tab styling */
    .stTabs [aria-selected="true"] {
        background-color: #3498db !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(52, 152, 219, 0.3);
        transform: translateY(-1px);
    }

    /* Tab content area */
    .stTabs [data-baseweb="tab-panel"] {
        background-color: #f8f9fa;
        border-radius: 0 0 10px 10px;
        padding: 20px;
        border: 1px solid #e9ecef;
        border-top: none;
        min-height: 500px;
    }

    /* Remove default border */
    .stTabs [data-baseweb="tab-border"] {
        display: none;
    }

    /* Custom tab icons (optional) */
    .stTabs [data-baseweb="tab"]:nth-child(1)::before {
        margin-right: 8px;
    }

    .stTabs [data-baseweb="tab"]:nth-child(2)::before {
        margin-right: 8px;
    }

    .stTabs [data-baseweb="tab"]:nth-child(3)::before {
        margin-right: 8px;
    }

    /* Main title styling */
    .main-title {
        text-align: center;
        color: #2c3e50;
        font-size: 2.5em;
        font-weight: bold;
        margin-bottom: 30px;
        padding: 20px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
</style>
""",
    unsafe_allow_html=True,
)


# =========================
# Document readers
# =========================
@st.cache_data
def read_docx(file_path):
    """Read DOCX file content"""
    try:
        doc = docx.Document(file_path)
        content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                content.append(paragraph.text)
        return "\n".join(content)
    except Exception as e:
        st.error(f"Error reading DOCX: {e}")
        return ""


@st.cache_data
def read_xlsx(file_path):
    """Read XLSX file content"""
    try:
        wb = load_workbook(file_path, read_only=True)
        content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content.append(f"Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    content.append(
                        " | ".join(
                            [str(cell) if cell is not None else "" for cell in row]
                        )
                    )
        return "\n".join(content)
    except Exception as e:
        st.error(f"Error reading XLSX: {e}")
        return ""


# =========================
# Persistent placeholders
# =========================
def get_progress_bar():
    if "_bar_slot" not in st.session_state:
        st.session_state["_bar_slot"] = st.empty()
        st.session_state["_bar"] = st.session_state["_bar_slot"].progress(0)
    return st.session_state["_bar"]


def reset_progress_bar():
    if "_bar_slot" in st.session_state:
        st.session_state["_bar_slot"].empty()
        del st.session_state["_bar_slot"]
    if "_bar" in st.session_state:
        del st.session_state["_bar"]


def ensure_flow_slot():
    if "_flow_slot" not in st.session_state:
        st.session_state["_flow_slot"] = st.empty()
    if "_flow_render_seq" not in st.session_state:
        st.session_state["_flow_render_seq"] = 0
    return st.session_state["_flow_slot"]


# =========================
# Stepper model + rendering
# =========================
def init_steps():
    st.session_state["_steps"] = [{"name": n, "status": "pending"} for n in STEP_NAMES]


def current_index_from_statuses() -> int:
    statuses = [s["status"] for s in st.session_state.get("_steps", [])]
    for i, s in enumerate(statuses):
        if s == "running":
            return i
    for i, s in enumerate(statuses):
        if s == "pending":
            return i
    return len(statuses) - 1 if statuses else 0


def render_stepper(steps, current_idx: int) -> go.Figure:
    n = len(steps)
    xs = list(range(n))
    ys = [0] * n
    GREEN = "#22C55E"
    BLUE = "#3B82F6"
    GRAY = "#CBD5E1"
    fig = go.Figure()
    fig.add_shape(
        type="line", x0=0, y0=0, x1=n - 1, y1=0, line=dict(color=GRAY, width=6)
    )
    if n > 1 and current_idx > 0:
        fig.add_shape(
            type="line",
            x0=0,
            y0=0,
            x1=current_idx,
            y1=0,
            line=dict(color=GREEN, width=6),
        )

    for i, label in enumerate(steps):
        color = GREEN if i < current_idx else BLUE if i == current_idx else GRAY
        fig.add_trace(
            go.Scatter(
                x=[xs[i]],
                y=[ys[i]],
                mode="markers+text",
                marker=dict(size=42, color=color, line=dict(width=2, color="white")),
                text=[label],
                textposition="top center",
                textfont=dict(size=12, color="#111827"),
                hoverinfo="text",
                showlegend=False,
            )
        )
        status = st.session_state["_steps"][i]["status"]
        fig.add_annotation(
            x=xs[i],
            y=-0.35,
            text=(
                "OK"
                if status == "ok"
                else "RUNNING ⏳" if status == "running" else "PENDING •"
            ),
            showarrow=False,
            font=dict(size=11, color="#111827"),
        )

    fig.update_layout(
        xaxis=dict(
            showgrid=False, zeroline=False, showticklabels=False, range=[-0.5, n - 0.5]
        ),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, range=[-1, 1]),
        height=220,
        margin=dict(l=16, r=16, t=24, b=12),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        title=dict(text="Assessment Progress", x=0.5, font=dict(size=18)),
    )
    return fig


def draw_stepper():
    """Draw/update the stepper with a UNIQUE key each time."""
    flow_slot = ensure_flow_slot()
    st.session_state["_flow_render_seq"] += 1
    key = f"flow_chart_{st.session_state['_flow_render_seq']}"
    flow_slot.plotly_chart(
        render_stepper(STEP_NAMES, current_index_from_statuses()),
        use_container_width=True,
        key=key,
    )


def set_step_status(step_name: str, status: str):
    for s in st.session_state["_steps"]:
        if s["name"] == step_name:
            s["status"] = status
            break
    draw_stepper()


# =========================
# Load models (cached)
# =========================
@st.cache_resource(show_spinner=False)
def load_models():
    reg = joblib.load(settings.REGRESSION_MODEL_PATH)
    clf = joblib.load(settings.SIGNAL_CLASSIFICATION_MODEL_PATH)
    scl = joblib.load(settings.STANDARD_SCALER_MODEL_PATH)
    return reg, clf, scl


# =========================
# Main App
# =========================
# Enhanced title with custom styling
st.markdown(
    '<h1 class="main-title">Credit Risk Assessment</h1>', unsafe_allow_html=True
)

# Create tabs with enhanced styling
tab1, tab2, tab3 = st.tabs(["Prediction", "Analysis", "Dashboard"])

with tab1:
    # Original prediction functionality
    ensure_flow_slot()
    if "_steps" not in st.session_state:
        init_steps()
    draw_stepper()

    if "_bar" not in st.session_state:
        get_progress_bar()

    try:
        reg_model, clf_model, scaler = load_models()
    except Exception as e:
        st.error(f"Model load failed: {e}")
        st.stop()

    uploaded_file = st.file_uploader(
        "Upload cleaned CSV (with Partner_Code)", type=["csv"]
    )

    if uploaded_file:
        # Fresh run: reset steps + progress bar + redraw once
        init_steps()
        reset_progress_bar()
        draw_stepper()
        bar = get_progress_bar()

        # Step 1: Load CSV
        try:
            set_step_status("Load CSV", "running")
            input_df = pd.read_csv(uploaded_file)
            id_cols = (
                input_df[["Partner_Code"]]
                if "Partner_Code" in input_df.columns
                else pd.DataFrame()
            )
            features_df = input_df.drop(columns=["Partner_Code"], errors="ignore")
            set_step_status("Load CSV", "ok")
            bar.progress(20)
        except Exception as e:
            set_step_status("Load CSV", "fail")
            st.error(f"CSV load failed: {e}")
            st.stop()

        # Step 2: Scale Features
        try:
            set_step_status("Scale Features", "running")
            features_df_num = features_df.apply(pd.to_numeric, errors="coerce")
            med = features_df_num.median(numeric_only=True)
            features_df_num = features_df_num.fillna(med).fillna(0)
            X_scaled = scaler.transform(features_df_num)
            set_step_status("Scale Features", "ok")
            bar.progress(40)
        except Exception as e:
            set_step_status("Scale Features", "fail")
            st.error(f" Scaling failed: {e}")
            st.stop()

        # Step 3: Predict Score
        try:
            set_step_status("Predict Score", "running")
            pred_score = reg_model.predict(X_scaled)
            set_step_status("Predict Score", "ok")
            bar.progress(60)
        except Exception as e:
            set_step_status("Predict Score", "fail")
            st.error(f" Score prediction failed: {e}")
            st.stop()

        # Step 4: Predict Signal
        try:
            set_step_status("Predict Signal", "running")
            pred_signal = clf_model.predict(X_scaled)
            signal_map = {0: "Green", 1: "Red", 2: "Yellow"}
            signal_labels = [signal_map.get(s, s) for s in pred_signal]
            set_step_status("Predict Signal", "ok")
            bar.progress(80)
        except Exception as e:
            set_step_status("Predict Signal", "fail")
            st.error(f" Signal prediction failed: {e}")
            st.stop()

        # Step 5: Build Output
        try:
            set_step_status("Build Output", "running")
            if "Partner_Code" not in input_df.columns:
                st.error(" Uploaded CSV must contain Partner_Code column.")
                st.stop()
            results = pd.DataFrame({"Partner_Code": input_df["Partner_Code"]})
            results["Predicted Score"] = pred_score
            results["Predicted Signal"] = signal_labels
            # === NEW: transactional enrichment ===
            from app.utils.cdp_utils import (
                aggregate_transactional_features,
                apply_final_scoring,
            )

            txn_df = pd.read_csv("partner_transactional_dataset.csv")
            agg_df = aggregate_transactional_features(
                txn_df, partner_col="Partner_Code"
            )

            final_results = apply_final_scoring(
                results_df=results, agg_df=agg_df, partner_key="Partner_Code", alpha=0.7
            )

            # Store enriched results in session state
            st.session_state["prediction_results"] = final_results
            st.session_state["input_features"] = agg_df

            set_step_status("Build Output", "ok")
            bar.progress(100)

            st.success(" Prediction Complete!")
            st.dataframe(
                final_results[
                    [
                        "Partner_Code",
                        "Predicted Score",
                        "Predicted Signal",
                        "Final Score",
                        "Final Signal",
                    ]
                ],
                use_container_width=True,
            )
            st.download_button(
                "Download Results",
                data=final_results.to_csv(index=False),
                file_name="cdp_predictions.csv",
                mime="text/csv",
            )
            st.session_state["cdp_input"] = input_df.copy()

        except Exception as e:
            set_step_status("Build Output", "fail")
            st.error(f"Building output failed: {e}")
    else:
        st.info("Upload a CSV to begin.")


# ------------------------------
# Tab 2: Signal Analysis & Reasoning
# ------------------------------
with tab2:
    st.header(" Signal Analysis")

    # Ensure predictions exist
    if (
        "prediction_results" not in st.session_state
        or st.session_state["prediction_results"] is None
    ):
        st.info("Please run predictions in the Prediction tab first.")
    else:
        results = st.session_state["prediction_results"]
        if results.empty:
            st.info("No prediction results found. Please run the Prediction tab.")
        else:
            # prefer final columns if available
            score_col = (
                "Final Score" if "Final Score" in results.columns else "Predicted Score"
            )
            signal_col = (
                "Final Signal"
                if "Final Signal" in results.columns
                else "Predicted Signal"
            )

            # partner selector
            partner_list = (
                results["Partner_Code"].tolist()
                if "Partner_Code" in results.columns
                else results.index.astype(str).tolist()
            )
            selected_partner = st.selectbox("Select Partner for Analysis", partner_list)

            # helper: robust partner row
            def _get_partner_row(results_df, partner_code):
                try:
                    return results_df[results_df["Partner_Code"] == partner_code].iloc[
                        0
                    ]
                except Exception:
                    try:
                        idx = results_df.index[
                            results_df.index.astype(str) == partner_code
                        ][0]
                        return results_df.loc[idx]
                    except Exception:
                        return results_df.iloc[0]

            # helper: get KPIs snapshot (from agg_df / input_features)
            def _get_partner_kpis(feats_df, partner_code):
                if feats_df is None:
                    return pd.Series(dtype="float64")
                try:
                    if "Partner_Code" in feats_df.columns:
                        tmp = feats_df[feats_df["Partner_Code"] == partner_code]
                        if not tmp.empty:
                            return tmp.iloc[0]
                    # fallback: align by ordinal position if possible
                    try:
                        ridx = results[results["Partner_Code"] == partner_code].index[0]
                        return feats_df.reset_index(drop=True).iloc[ridx]
                    except Exception:
                        return pd.Series(dtype="float64")
                except Exception:
                    return pd.Series(dtype="float64")

            # helper: get original CDP input row (if stored at Tab1)
            def _get_model_input_row(partner_code):
                if "cdp_input" not in st.session_state:
                    return pd.Series(dtype="float64")
                df_input = st.session_state["cdp_input"]
                try:
                    if "Partner_Code" in df_input.columns:
                        tmp = df_input[df_input["Partner_Code"] == partner_code]
                        if not tmp.empty:
                            return tmp.drop(
                                columns=["Partner_Code"], errors="ignore"
                            ).iloc[0]
                except Exception:
                    pass
                return pd.Series(dtype="float64")

            # Analyze single partner
            if st.button("Analyze Signal Reasoning"):
                with st.spinner("Analyzing signal reasoning..."):
                    try:
                        prow = _get_partner_row(results, selected_partner)
                        kpis = _get_partner_kpis(
                            st.session_state.get("input_features"), selected_partner
                        )
                        model_input = _get_model_input_row(selected_partner)

                        # build KPI text (select important KPI keys)
                        candidate_kpis = [
                            "avg_days_past_due",
                            "max_days_past_due",
                            "late_ratio",
                            "dispute_ratio",
                            "adjustment_ratio",
                            "allocation_ratio",
                            "collection_rate",
                            "total_invoices",
                            "total_installments",
                            "total_invoice_amount",
                        ]
                        kpi_lines = []
                        for k in candidate_kpis:
                            if k in kpis.index:
                                v = kpis.get(k)
                                kpi_lines.append(f"- {k}: {v}")
                        kpi_text = (
                            "\n".join(kpi_lines)
                            if kpi_lines
                            else "(no transactional KPI snapshot available)"
                        )

                        # build input-feature text: top numeric features only (max 6)
                        input_text = "(no original input snapshot available)"
                        if isinstance(model_input, pd.Series) and not model_input.empty:
                            try:
                                numeric_input = pd.to_numeric(
                                    model_input, errors="coerce"
                                ).dropna()
                                if not numeric_input.empty:
                                    top_input = numeric_input.abs().nlargest(6)
                                    input_lines = [
                                        f"- {nm}: {val:.3f}"
                                        for nm, val in top_input.items()
                                    ]
                                    input_text = "\n".join(input_lines)
                            except Exception:
                                input_text = "(unable to render input features)"

                        # prepare summary numbers for UI
                        st.subheader("Partner Summary")
                        st.metric("Partner Code", selected_partner)
                        fs_val = prow.get(
                            "Final Score", prow.get("Predicted Score", np.nan)
                        )
                        st.metric(
                            "Final Score",
                            f"{fs_val:.2f}" if pd.notna(fs_val) else "N/A",
                        )
                        sig_val = prow.get(
                            "Final Signal", prow.get("Predicted Signal", "N/A")
                        )
                        st.metric("Final Signal", sig_val)

                        # Build the AI prompt (concise, includes both KPIs and model input snapshot)
                        final_score_str = f"{fs_val:.2f}" if pd.notna(fs_val) else "N/A"
                        final_signal_str = sig_val
                        prompt = f"""
You are a senior credit risk analyst. Provide a structured, actionable analysis for Partner {selected_partner}.

Context:
- Final Score: {final_score_str}
- Final Signal: {final_signal_str}
- Model Predicted Score: {prow.get('Predicted Score', 'N/A')}
- Model Predicted Signal: {prow.get('Predicted Signal', 'N/A')}
- Override reason: {prow.get('OverrideReason', 'None')}
- Adjustment delta: {prow.get('AdjustmentDelta', 'N/A')}

Partner Transactional KPIs (recent performance):
{kpi_text}

Partner CDP Input Features (financial & operational data):
{input_text}

Guidelines:
- Always combine insights from BOTH Transactional KPIs AND CDP Input Features.
- Highlight any contradictions (e.g., strong KPIs but weak financial ratios, or vice versa).
- Mention differences between Model Predicted and Final Score/Signal.

Requirements:
1) Key Factors (3 bullets) — must cover both transactional behavior and CDP financial indicators.
2) Risk Assessment (1–2 sentences) — balanced view from both data sources.
3) Recommendation (approve / review / decline) with justification.
4) Quick Action Plan (3 concise steps).
Keep under 250 words and
 Make sure the confidentiality maintain use the data whatever the data is needed for analysis but try not to reveal the data numbers from dataset in the response .
"""

                        # Call AI if available, else show friendly message
                        if not settings.is_gemini_available:
                            ai_text = "AI not configured. Please set GEMINI_API_KEY in your environment or paste key in session."
                        else:
                            try:
                                ai_text = analyze_with_gemini(prompt)
                            except Exception as e:
                                ai_text = f"Error calling AI: {e}"

                        st.subheader("AI Analysis")
                        st.markdown(ai_text)

                    except Exception as e:
                        st.error(f"Error while preparing analysis: {e}")

            # Batch analysis (safe, uses same KPI prompt format)
            st.markdown("---")
            st.subheader(" Batch Analysis")
            confirm_batch = st.checkbox(
                "I understand this will call the AI for each partner and may incur costs."
            )
            run_batch = st.button("Analyze All Red/Yellow Signals")
            if run_batch:
                if not confirm_batch:
                    st.info(
                        "Please check the confirmation checkbox before running batch analysis."
                    )
                else:
                    signal_col_local = (
                        "Final Signal"
                        if "Final Signal" in results.columns
                        else "Predicted Signal"
                    )
                    score_col_local = (
                        "Final Score"
                        if "Final Score" in results.columns
                        else "Predicted Score"
                    )
                    high_risk = results[
                        results[signal_col_local].isin(["Red", "Yellow"])
                    ]
                    if high_risk.empty:
                        st.info("No Red/Yellow partners found.")
                    else:
                        progress = st.progress(0)
                        out = []
                        for i, (_, partner) in enumerate(high_risk.iterrows()):
                            try:
                                partner_code = partner.get(
                                    "Partner_Code", str(partner.name)
                                )
                                prow = _get_partner_row(results, partner_code)
                                kpis = _get_partner_kpis(
                                    st.session_state.get("input_features"), partner_code
                                )
                                # small KPI text for batch
                                kp_items = []
                                for k in [
                                    "late_ratio",
                                    "avg_days_past_due",
                                    "max_days_past_due",
                                    "dispute_ratio",
                                    "collection_rate",
                                ]:
                                    if k in kpis.index:
                                        kp_items.append(f"- {k}: {kpis.get(k)}")
                                kp_text_batch = (
                                    "\n".join(kp_items) if kp_items else "(no kpi)"
                                )
                                prompt_batch = f"Partner {partner_code} | Signal: {prow.get(signal_col_local)} | Score: {prow.get(score_col_local)}\\nTop KPIs:\\n{kp_text_batch}\\nProvide 3 bullets: primary risk, secondary concern, recommended action (each <=25 words)."
                                ai_resp = (
                                    analyze_with_gemini(prompt_batch)
                                    if settings.is_gemini_available
                                    else "AI not configured"
                                )
                                out.append(
                                    {
                                        "Partner_Code": partner_code,
                                        "Signal": prow.get(signal_col_local),
                                        "Score": prow.get(score_col_local),
                                        "Analysis": ai_resp,
                                    }
                                )
                            except Exception as e:
                                out.append(
                                    {
                                        "Partner_Code": partner.get(
                                            "Partner_Code", str(partner.name)
                                        ),
                                        "Signal": partner.get(signal_col_local),
                                        "Score": partner.get(score_col_local),
                                        "Analysis": f"Error: {e}",
                                    }
                                )
                            progress.progress((i + 1) / len(high_risk))
                        st.dataframe(pd.DataFrame(out), use_container_width=True)


with tab3:
    render_dashboard(
        prediction_results=st.session_state.get("prediction_results"),  # preferred
        input_features=st.session_state.get("input_features"),  # optional
        analyze_fn=analyze_with_gemini,  # your existing analyze function
        gemini_model_name="gemini-2.5-flash",
        use_upload=False,  # <- important: no uploader shown in Tab3
    )
